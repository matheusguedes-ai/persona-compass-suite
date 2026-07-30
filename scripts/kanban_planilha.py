#!/usr/bin/env python3
"""
Gera a planilha de gestao das demandas (Kanban) a partir de scripts/kanban_dados.json.

Por que um script e nao a planilha direto: a fonte da verdade fica NO REPO,
versionada. Assim da para ver no git o que mudou de status e quando, e a
planilha pode ser regerada identica se alguem apagar a do Drive.

Saida: kanban-demandas.xlsx, que o conector do Drive converte para Google
Sheets na hora do upload (abas, cores, listas suspensas e formulas sobrevivem).

Estrutura:
  - "Demandas"   -> a fonte: uma linha por demanda, com lista suspensa no Status
  - "Kanban"     -> o quadro: quatro colunas que se remontam sozinhas por FILTER
  - "Como usar"  -> as regras, escritas na propria planilha

Uso: python3 scripts/kanban_planilha.py
"""
import json
import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

AQUI = os.path.dirname(os.path.abspath(__file__))
DADOS = os.path.join(AQUI, "kanban_dados.json")
SAIDA = os.path.join(os.path.dirname(AQUI), "kanban-demandas.xlsx")

# As quatro colunas que ele pediu, na ordem em que uma demanda anda.
COLUNAS_KANBAN = ["Planejado", "Em andamento", "Concluído", "Precisa de melhorias"]

COR = {
    "Planejado": "E8EAF0",
    "Em andamento": "FFF0CC",
    "Concluído": "D9F0DC",
    "Precisa de melhorias": "FBDDD8",
}
COR_TEXTO = {
    "Planejado": "3C4257",
    "Em andamento": "7A5B00",
    "Concluído": "1B5E20",
    "Precisa de melhorias": "8C2A1C",
}

CABECALHO = [
    ("#", 5),
    ("Demanda", 42),
    ("O que é", 62),
    ("Área", 14),
    ("Status", 20),
    ("Prioridade", 11),
    ("De quem depende", 16),
    ("Origem do pedido", 30),
    ("Prova / ressalva", 52),
]

FILL_TITULO = PatternFill("solid", fgColor="1F2937")
FONT_TITULO = Font(color="FFFFFF", bold=True, size=11)
BORDA = Border(*[Side(style="thin", color="D0D5DD")] * 4)


def carregar():
    with open(DADOS, encoding="utf-8") as f:
        d = json.load(f)
    itens = d["demandas"]
    assert itens, "kanban_dados.json sem demandas"
    for i, it in enumerate(itens, 1):
        faltando = {"titulo", "descricao", "area", "status", "prioridade",
                    "depende_de", "origem", "prova"} - set(it)
        assert not faltando, f"demanda {i} sem os campos {faltando}"
        assert it["status"] in COLUNAS_KANBAN, \
            f"demanda {i} com status fora do quadro: {it['status']!r}"
    # Prioridade unica e contigua: dois "1" fazem a ordenacao mentir.
    prios = sorted(it["prioridade"] for it in itens)
    assert prios == list(range(1, len(itens) + 1)), \
        f"prioridades precisam ser 1..{len(itens)} sem repetir; vieram {prios}"
    return sorted(itens, key=lambda x: x["prioridade"])


def aba_demandas(wb, itens):
    ws = wb.create_sheet("Demandas")
    ws.append([c[0] for c in CABECALHO])
    for i, (_, larg) in enumerate(CABECALHO, 1):
        ws.column_dimensions[get_column_letter(i)].width = larg
        c = ws.cell(row=1, column=i)
        c.fill, c.font = FILL_TITULO, FONT_TITULO
        c.alignment = Alignment(vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 26

    for it in itens:
        ws.append([
            it["prioridade"], it["titulo"], it["descricao"], it["area"],
            it["status"], it["prioridade"], it["depende_de"],
            it["origem"], it["prova"],
        ])

    ultima = len(itens) + 1
    for linha in ws.iter_rows(min_row=2, max_row=ultima, max_col=len(CABECALHO)):
        for c in linha:
            c.border = BORDA
            c.alignment = Alignment(vertical="top", wrap_text=True)

    # Lista suspensa no Status: digitar "concluido" a mao quebraria o quadro,
    # que casa o texto exato.
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(COLUNAS_KANBAN) + '"',
        allow_blank=False,
        showDropDown=False,
    )
    dv.error = "Use um dos quatro status do quadro."
    dv.errorTitle = "Status inválido"
    ws.add_data_validation(dv)
    dv.add(f"E2:E{max(ultima, 200)}")

    # Cor por status na linha inteira: o olho acha o vermelho antes de ler.
    faixa = f"A2:I{max(ultima, 200)}"
    for status in COLUNAS_KANBAN:
        ws.conditional_formatting.add(
            faixa,
            FormulaRule(
                formula=[f'$E2="{status}"'],
                fill=PatternFill("solid", bgColor=COR[status]),
                stopIfTrue=False,
            ),
        )

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:I{ultima}"
    return ws


def aba_kanban(wb, total):
    """O quadro. Nao guarda dado: le a aba Demandas por formula.

    Assim mudar a lista suspensa de uma linha move o cartao de coluna sozinho —
    que e o mais perto de "tempo real" que uma planilha chega sem robo.
    """
    ws = wb.create_sheet("Kanban", 0)
    fim = max(total + 1, 300)

    ws["A1"] = "Quadro de demandas — Thrive Profiler"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = ("Este quadro se monta sozinho a partir da aba Demandas. "
                "Para mover um cartão, mude o Status lá — não digite aqui.")
    ws["A2"].font = Font(size=10, italic=True, color="667085")

    for i, status in enumerate(COLUNAS_KANBAN):
        col = get_column_letter(i * 2 + 1)
        ws.column_dimensions[col].width = 46
        ws.column_dimensions[get_column_letter(i * 2 + 2)].width = 3

        t = ws[f"{col}4"]
        t.value = f'="{status}  ("&COUNTIF(Demandas!$E:$E,"{status}")&")"'
        t.fill = PatternFill("solid", fgColor=COR[status])
        t.font = Font(bold=True, size=12, color=COR_TEXTO[status])
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[4].height = 30

        # Um FILTER por coluna: a lista inteira sai de uma formula so, e uma
        # demanda nova na aba Demandas aparece aqui sem ninguem mexer no quadro.
        ws[f"{col}5"] = (
            f'=IFERROR(SORT(FILTER(Demandas!$B$2:$B${fim}&" · "&Demandas!$D$2:$D${fim},'
            f'Demandas!$E$2:$E${fim}="{status}"),1,TRUE),"—")'
        )
        ws[f"{col}5"].alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A5"
    return ws


def aba_como_usar(wb, itens):
    ws = wb.create_sheet("Como usar")
    ws.column_dimensions["A"].width = 110
    linhas = [
        ("Como esta planilha funciona", True),
        ("", False),
        ("A aba Demandas é a fonte. A aba Kanban é só uma vista: ela se remonta sozinha", False),
        ("quando você muda o Status. Não digite nada direto no quadro — some na próxima", False),
        ("atualização.", False),
        ("", False),
        ("Os quatro status", True),
        ("Planejado — pedido registrado, ainda não começou.", False),
        ("Em andamento — está sendo construído agora.", False),
        ("Concluído — no ar e conferido na plataforma.", False),
        ("Precisa de melhorias — entregue, mas com ressalva: algo pela metade, um bug", False),
        ("conhecido, ou não testado com usuário real. A coluna 'Prova / ressalva' diz o quê.", False),
        ("", False),
        ("De quem depende", True),
        ("Claude — é meu, eu faço.", False),
        ("Matheus — decisão sua, senha sua, ou configuração fora da plataforma.", False),
        ("Depende de dados — só faz sentido quando houver amostra real na base.", False),
        ("", False),
        ("Quem atualiza", True),
        ("Eu atualizo esta planilha sempre que uma demanda mudar de estado no nosso", False),
        ("trabalho, e você pode pedir 'atualize o kanban' a qualquer momento. Eu não rodo", False),
        ("sozinho o dia inteiro — então 'tempo real' aqui quer dizer: a cada sessão em que", False),
        ("mexermos em algo, e sempre no fim de uma entrega.", False),
        ("", False),
        ("Você pode editar à vontade: mover status, mudar prioridade, escrever nas células.", False),
        ("Eu leio a planilha antes de atualizar, então o que você escrever não se perde.", False),
        ("", False),
        (f"Gerada em {date.today().strftime('%d/%m/%Y')} · {len(itens)} demandas · "
         "fonte no repo: scripts/kanban_dados.json", False),
    ]
    for texto, titulo in linhas:
        ws.append([texto])
        c = ws.cell(row=ws.max_row, column=1)
        c.font = Font(bold=True, size=12) if titulo else Font(size=10)
        c.alignment = Alignment(wrap_text=False, vertical="center")
    return ws


def main():
    itens = carregar()
    wb = Workbook()
    wb.remove(wb.active)
    aba_demandas(wb, itens)
    aba_kanban(wb, len(itens))
    aba_como_usar(wb, itens)
    wb.move_sheet("Kanban", offset=-2)
    wb.save(SAIDA)

    por_status = {s: sum(1 for i in itens if i["status"] == s) for s in COLUNAS_KANBAN}
    print(f"{SAIDA}")
    print(f"{len(itens)} demandas: " + " · ".join(f"{s} {n}" for s, n in por_status.items()))


if __name__ == "__main__":
    main()
